import openpyxl


wb = openpyxl.load_workbook('example.xlsx')
ws = wb['student1']

# 按行获取表格的内容，返回是迭代器,注意这里返回的是数据，不是Cell对象
for row in ws.values:
    # 内部每一行的数据以元组的形式展现
    print(*row)

# 根据rows属性来遍历表格中的数据，类似iter_rows,返回的每行的cell对象
for row in ws.rows:
    print([cell.value for cell in row])

for row in ws.iter_rows():
    print([cell.value for cell in row])