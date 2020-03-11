"""
excel文件合并的大致内容
1.找出目录下的所有excel文件，一般根据后缀结尾匹配
2.读取所有找到的excel文件中的第一个文件，作为起始文件
3.逐一读取后面的所有文件，从第二行开始遍历数据，将遍历出来的数据添加到起始文件中
4.保存起始文件即可
"""

import os
import glob
import openpyxl


def get_all_xlsx_file(path):
    """
    找出目录下的所有xlsx文件
    :param path: 目录
    :return: 得到所有的xlsx文件名称
    """
    # glob匹配出所有xlsx结尾的文件
    xlsx_files = glob.glob(os.path.join(path, '*.xlsx'))
    # 排序
    sorted(xlsx_files, key=str.lower)
    return xlsx_files


def merge_xlsx_files(xlsx_files):
    # 先打开第一个文档，作为初始文档，其他文件的内容往里添加即可
    wb = openpyxl.load_workbook(xlsx_files[0])
    ws = wb.active
    ws.title = 'Merged result'

    for file in xlsx_files[1:]:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            # 通过append函数将数据添加到汇总表的末尾
            ws.append(values)
    return wb


def main():
    xlsx_files = get_all_xlsx_file('excel_files')
    wb = merge_xlsx_files(xlsx_files)
    wb.save('merged_form.xlsx')


if __name__ == '__main__':
    main()
