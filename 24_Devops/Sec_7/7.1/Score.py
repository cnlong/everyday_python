import openpyxl


def process_worksheet(sheet):
    # 找出最大列，之后两列分别作为总分列和平均分列
    avg_column = sheet.max_column + 1
    sum_column = sheet.max_column + 2



    # 遍历列表，读取每行的数据，注意，第一列保存的数据是学号，第二列保存的是学生的姓名，所以从第三列开始读取
    # 第一行是表头，所以从第二行开始读取，所以使用iter_rows遍历的时候从第二行第三列遍历
    # 遍历每行的数据
    for row in sheet.iter_rows(min_row=2, min_col=3):
        # 找出每个单元格的数据
        scores = [cell.value for cell in row]
        # 计算总分和平均分
        sum_score = sum(scores)
        avg_score = sum_score / len(scores)

        # 保存数据
        sheet.cell(row=row[0].row, column=avg_column).value = avg_score
        sheet.cell(row=row[0].row, column=sum_column).value = sum_score

    # 这一步必须写在后面，不然会将这下面的空字符串作为上面行的数据计算，会报错
    sheet.cell(row=1, column=avg_column).value = 'Avg'
    sheet.cell(row=1, column=sum_column).value = 'Sum'


def main():
    wb = openpyxl.load_workbook('example2.xlsx')
    sheet = wb['Student']
    process_worksheet(sheet)
    wb.save('example2_copy.xlsx')


if __name__ == '__main__':
    main()
