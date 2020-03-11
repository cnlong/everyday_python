from openpyxl import Workbook
import datetime

# 新建一个Workbook
wb = Workbook()

# 获取当前活跃的sheet
ws = wb.active
# 也可以通过以下方法获取一个ws
# print(wb['Sheet'])
# 获取sheet标题
print(ws.title)
# 修改属性值
ws.title = 'Student'
print(ws.title)

# 创建新的表格（sheet）
wb.create_sheet(index=0, title='new sheet')
print(wb.sheetnames)

# 删除表格
wb.remove(wb['Student'])
print(wb.sheetnames)

ws = wb.active
print(ws.title)

# 填充数据
ws['A1'] = "hello, world"
ws['A2'] = datetime.datetime.now()

# 保存文件
wb.save('sample.xlsx')