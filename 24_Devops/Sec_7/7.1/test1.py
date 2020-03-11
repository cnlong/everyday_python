import openpyxl

# 读取一个excel文档
wb = openpyxl.load_workbook('example.xlsx')
# 或者直接创建一个excel文档，向里面写入内容
# wb = openpyxl.Workbook()

# 获取活跃的worksheet
print(wb.active)

# 是否以只读模式打开Excel文档
print(wb.read_only)

# 文档的字符集编码
print(wb.encoding)

# 列表形式返回所有的Worksheet
print(wb.worksheets)

# 获取所有表格的名称
print(wb.sheetnames)

# 通过表格名称获取表格
print(wb[u'student1'])

# 获取某一个sheet
ws = wb['student1']
# 获取表格的标题
print(ws.title)
# 获取表格大小，从左上角到右下角表格的坐标
print(ws.dimensions)
# 获取表格的最大列
print(ws.max_column)
# 表格的最小列
print(ws.min_column)

# 获取表格的最大行
print(ws.max_row)
# 表格的最小行
print(ws.min_row)
# 按行获取单元格，返回的是一个Cell对象
print(ws.rows)
# 按列获取单元格，返回的是一个Cell对象
print(ws.columns)
# 按行获取表格的内容，返回是迭代器,注意这里返回的是数据，不是Cell对象
print(ws.values)

# 按行获取所有的单元格,后面的参数指定获取的范围
print(list(ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3)))

# 大多数方法获取返回的是Cell对象，一个对象代表一个单元格，也可以直接通过坐标的方式获取Cell对象，或者通过cell方法指定行列获取Cell对象
print(ws['A2'])
print(ws.cell(row=2, column=1))

# 获取单元格所在的行
print(ws['A2'].row)

# 获取单元格所在的列
print(ws['A2'].column)

# 获取单元格所在的取值
print(ws['A2'].value)


